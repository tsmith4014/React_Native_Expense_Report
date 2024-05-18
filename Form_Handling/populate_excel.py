#populate_excel.py
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from datetime import datetime, timedelta
import os

def populate_template(data, template_path, output_path):
    try:
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active

        img_path = os.path.join(os.path.dirname(__file__), 'eahead.jpg')
        pil_img = PILImage.open(img_path)
        img = Image(pil_img)
        img.width, img.height = img.width * 0.43, img.height * 0.60
        sheet.add_image(img, 'A1')

        center_aligned_text = Alignment(horizontal='center', wrapText=True)

        period_ending_date = datetime.strptime(data['period_ending'], '%Y-%m-%d')
        period_start_date = period_ending_date - timedelta(days=6)

        cells_to_populate = {
            'B5': data['school'],
            'H4': period_ending_date.date(),
            'H5': data['trip_purpose'],
            'B4': data['employee_department']  # Employee/Department
        }

        for cell_ref, value in cells_to_populate.items():
            cell = sheet[cell_ref]
            cell.alignment = center_aligned_text
            cell.value = value

        day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        for i, day_name in enumerate(day_names):
            day_cell = sheet.cell(row=7, column=4+i)
            day_cell.value = f"Date\n{day_name}"
            day_cell.alignment = Alignment(horizontal='center', wrapText=True)

        for i in range(6, -1, -1):
            date_cell = sheet.cell(row=8, column=4+i)
            date_cell.value = (period_ending_date - timedelta(days=6-i)).date()
            date_cell.alignment = Alignment(horizontal='center', wrapText=True)

        file_data_by_date = {}
        for file_data in data['files_data']:
            file_date = file_data['date']
            if file_date not in file_data_by_date:
                file_data_by_date[file_date] = []
            file_data_by_date[file_date].append(file_data)

        for i in range(4, 11):
            date_cell = sheet.cell(row=8, column=i).value
            if date_cell in file_data_by_date:
                for file_data in file_data_by_date[date_cell]:
                    category_row = get_category_row(file_data['category'])
                    sheet.cell(row=category_row, column=i).value = file_data['price']

        # Handling travel dates for per diem fields
        if data.get('travel').lower() == 'yes' and all(data.get(key) for key in ['travel_start_date', 'travel_end_date']):
            populate_travel_dates(sheet, data['travel_start_date'], data['travel_end_date'])

    except Exception as e:
        print(f"An error occurred while populating the template: {e}")
        raise

    wb.save(output_path)
    return output_path

def get_category_row(category):
    category_mapping = {
        'Airfare': 10,
        'Car Rental': 11,
        'Local Transportation': 12,
        'Tolls/Parking': 13,
        'Car Expense': 14,
        'Gas': 15,
        'Hotel': 16,
        'Telephone': 17,
        'Business Meals': 18,
        'Entertainment': 19,
        'Office Supplies': 20,
        'Postage': 21,
        'Tips': 22,
        'Other': 23
    }
    return category_mapping.get(category, 24)

def populate_travel_dates(sheet, travel_start_date_str, travel_end_date_str):
    """Handles the logic for populating travel dates and per diem fields."""
    travel_start_date = datetime.strptime(travel_start_date_str, '%Y-%m-%d')
    travel_end_date = datetime.strptime(travel_end_date_str, '%Y-%m-%d')
    
    for i in range(4, 11):  # Assuming date columns are D to J (4 to 10)
        cell_date = sheet.cell(row=8, column=i).value
        if cell_date and travel_start_date.date() <= cell_date <= travel_end_date.date():
            if cell_date == travel_start_date.date():
                sheet.cell(row=21, column=i).value = 30.00  # Only dinner on start date
            elif cell_date == travel_end_date.date():
                sheet.cell(row=19, column=i).value = 5.00  # Only breakfast on end date
            else:
                sheet.cell(row=19, column=i).value = 5.00  # Breakfast
                sheet.cell(row=21, column=i).value = 30.00  # Dinner








#populate_excel.py
# import openpyxl
# from openpyxl.styles import Alignment
# from openpyxl.drawing.image import Image
# from datetime import datetime, timedelta
# import os

# def populate_template(data, template_path, output_path):
#     """
#     Populates an Excel template with data and saves it to a new file.
#     Incorporates logic for handling travel dates to determine per diem fields.
#     Includes error handling for robustness.
#     """
#     try:
#         # Load the Excel template
#         wb = openpyxl.load_workbook(template_path)
#         sheet = wb.active

#         # Insert and resize the image
#         # img_path = os.path.join(os.path.dirname(__file__), 'eahead.jpg')
#         # img = Image(img_path)
#         # img.width, img.height = img.width * 0.43, img.height * 0.60
#         # sheet.add_image(img, 'A1')

#         # Common setup for cell alignment
#         center_aligned_text = Alignment(horizontal='center', wrapText=True)

#         # Populate cells with provided data
#         cells_to_populate = {
#             'B5': data['school'],
#             'H4': datetime.strptime(data['period_ending'], '%Y-%m-%d').date(),
#             'H5': data['trip_purpose'],
#             'B4': data['employee_department']  # Employee/Department
#         }

#         for cell_ref, value in cells_to_populate.items():
#             cell = sheet[cell_ref]
#             cell.alignment = center_aligned_text
#             cell.value = value

#         # Handling date names and values
#         populate_dates(sheet, data['period_ending'])

#         # Handling travel dates for per diem fields
#         if 'yes' == data.get('travel') and all(data.get(key) for key in ['travel_start_date', 'travel_end_date']):
#             populate_travel_dates(sheet, data['travel_start_date'], data['travel_end_date'])

#     except Exception as e:
#         # Log the error or handle it as needed
#         print(f"An error occurred while populating the template: {e}")
#         raise

#     # Save the populated template to a new file
#     wb.save(output_path)

#     # Return the output path
#     return output_path



# def populate_dates(sheet, period_ending):
#     """Populates static day names and dates."""
#     period_ending_date = datetime.strptime(period_ending, '%Y-%m-%d')
#     day_names = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
#     for i, day_name in enumerate(day_names):
#         day_cell = sheet.cell(row=7, column=4+i)
#         day_cell.value = f"Date\n{day_name}"
#         day_cell.alignment = Alignment(horizontal='center', wrapText=True)

#     for i in range(6, -1, -1):
#         date_cell = sheet.cell(row=8, column=4+i)
#         date_cell.value = (period_ending_date - timedelta(days=6-i)).date()
#         date_cell.alignment = Alignment(horizontal='center', wrapText=True)

# def populate_travel_dates(sheet, travel_start_date_str, travel_end_date_str):
#     """Handles the logic for populating travel dates and per diem fields."""
#     travel_start_date = datetime.strptime(travel_start_date_str, '%Y-%m-%d')
#     travel_end_date = datetime.strptime(travel_end_date_str, '%Y-%m-%d')
    
#     for i in range(4, 11):  # Assuming date columns are D to J (4 to 10)
#         cell_date = sheet.cell(row=8, column=i).value
#         if cell_date and travel_start_date.date() <= cell_date <= travel_end_date.date():
#             if cell_date == travel_start_date.date():
#                 sheet.cell(row=21, column=i).value = 30.00  # Only dinner on start date
#             elif cell_date == travel_end_date.date():
#                 sheet.cell(row=19, column=i).value = 5.00  # Only breakfast on end date
#             else:
#                 sheet.cell(row=19, column=i).value = 5.00  # Breakfast
#                 sheet.cell(row=21, column=i).value = 30.00  # Dinner

